# Copyright (c) 2025, Abbass Chokor and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt, getdate, nowdate, add_to_date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class HPAmplify(Document):
	pass


@frappe.whitelist()
def export_hp_amplify_report(from_date, to_date, warehouses, suppliers, brands=None, item_groups=None, reporter_id=None, items_force_add=None, items_force_remove=None):
	"""
	Generate Excel report for HP Amplify with item-wise sales and stock data
	"""
	# Validate dates
	if not from_date or not to_date:
		frappe.throw(_("From Date and To Date are required"))
	
	from_date = getdate(from_date)
	to_date = getdate(to_date)
	
	if from_date > to_date:
		frappe.throw(_("From Date cannot be greater than To Date"))
	
	# Validate reporter_id
	if not reporter_id:
		frappe.throw(_("Reporter ID is required"))
	
	# Parse JSON data if needed
	import json
	if isinstance(warehouses, str):
		warehouses = json.loads(warehouses)
	if isinstance(suppliers, str):
		suppliers = json.loads(suppliers)
	if isinstance(brands, str):
		brands = json.loads(brands) if brands else []
	if isinstance(item_groups, str):
		item_groups = json.loads(item_groups) if item_groups else []
	if isinstance(items_force_add, str):
		items_force_add = json.loads(items_force_add) if items_force_add else []
	if isinstance(items_force_remove, str):
		items_force_remove = json.loads(items_force_remove) if items_force_remove else []
	
	# Build warehouse mapping (warehouse -> amplify_code)
	warehouse_mapping = {}
	for w in warehouses:
		warehouse = w.get('warehouse')
		amplify_code = w.get('warehouse_amplify_code') or warehouse
		warehouse_mapping[warehouse] = amplify_code
	
	# Get report data
	data = get_hp_amplify_data(from_date, to_date, warehouses, suppliers, brands, item_groups, warehouse_mapping, items_force_add, items_force_remove)
	
	# Generate Excel file
	file_path = generate_excel_report(data, from_date, to_date, reporter_id)
	
	return file_path


def get_hp_amplify_data(from_date, to_date, warehouses, suppliers, brands, item_groups, warehouse_mapping, items_force_add, items_force_remove):
	"""
	Get consolidated data for HP Amplify report
	"""
	# Extract warehouse and supplier lists
	warehouse_list = [w.get('warehouse') for w in warehouses if w.get('warehouse')]
	supplier_list = [s.get('supplier') for s in suppliers if s.get('supplier')]
	brand_list = [b.get('brand') for b in brands if b.get('brand')] if brands else []
	item_group_list = [ig.get('item_group') for ig in item_groups if ig.get('item_group')] if item_groups else []
	force_add_list = [i.get('item') for i in items_force_add if i.get('item')] if items_force_add else []
	force_remove_list = [i.get('item') for i in items_force_remove if i.get('item')] if items_force_remove else []
	
	if not warehouse_list:
		frappe.throw(_("At least one warehouse must be selected"))
	
	# Build warehouse hierarchy (parent -> children mapping)
	warehouse_hierarchy = build_warehouse_hierarchy(warehouse_list)
	
	# Get items based on filters
	items = get_filtered_items(warehouse_list, supplier_list, brand_list, item_group_list, force_add_list, force_remove_list)
	
	# Build report data
	report_data = []
	
	for item in items:
		item_code = item.get('item_code')
		has_serial_no = item.get('has_serial_no')
		
		# Check if item needs serial number tracking
		if has_serial_no and supplier_list:
			# Track by serial number for supplier-specific items
			serial_data = get_serial_number_data(item_code, warehouse_list, supplier_list, from_date, to_date, warehouse_hierarchy, warehouse_mapping)
			report_data.extend(serial_data)
		else:
			# Regular item tracking without serial numbers
			item_data = get_item_warehouse_data(item_code, warehouse_list, from_date, to_date, supplier_list, warehouse_hierarchy, warehouse_mapping)
			report_data.extend(item_data)
	
	return report_data


def get_filtered_items(warehouse_list, supplier_list, brand_list, item_group_list, force_add_list, force_remove_list):
	"""
	Get items filtered by brands, item groups, and suppliers
	Item groups include children if specified
	Force remove: Exclude even if matches conditions
	Force add: Include even if doesn't match conditions
	"""
	# Build item filter conditions
	conditions = ["i.disabled = 0"]
	
	if brand_list:
		brand_conditions = " OR ".join([f"i.brand = '{brand}'" for brand in brand_list])
		conditions.append(f"({brand_conditions})")
	
	# Handle item groups with children
	if item_group_list:
		# Get all item groups including children
		all_item_groups = get_item_groups_with_children(item_group_list)
		item_group_conditions = " OR ".join([f"i.item_group = '{ig}'" for ig in all_item_groups])
		conditions.append(f"({item_group_conditions})")
	
	# Get items that have stock in the warehouses or were purchased from suppliers
	query = f"""
		SELECT DISTINCT 
			i.item_code,
			i.item_name,
			i.has_serial_no,
			i.brand,
			i.item_group
		FROM `tabItem` i
		WHERE {' AND '.join(conditions)}
		AND (
			EXISTS (
				SELECT 1 FROM `tabBin` b 
				WHERE b.item_code = i.item_code 
				AND b.warehouse IN ({', '.join(['%s'] * len(warehouse_list))})
			)
			OR EXISTS (
				SELECT 1 FROM `tabPurchase Order Item` poi
				INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
				WHERE poi.item_code = i.item_code 
				AND po.docstatus = 1
				{f"AND po.supplier IN ({', '.join(['%s'] * len(supplier_list))})" if supplier_list else ""}
			)
			OR EXISTS (
				SELECT 1 FROM `tabPurchase Receipt Item` pri
				INNER JOIN `tabPurchase Receipt` pr ON pr.name = pri.parent
				WHERE pri.item_code = i.item_code 
				AND pr.docstatus = 1
				{f"AND pr.supplier IN ({', '.join(['%s'] * len(supplier_list))})" if supplier_list else ""}
			)
		)
		ORDER BY i.item_code
	"""
	
	params = warehouse_list.copy()
	if supplier_list:
		params.extend(supplier_list)
		params.extend(supplier_list)
	
	items = frappe.db.sql(query, params, as_dict=1)
	
	# Apply force remove - Remove items that are in force_remove_list
	if force_remove_list:
		items = [item for item in items if item.get('item_code') not in force_remove_list]
	
	# Apply force add - Add items from force_add_list that are not already included
	if force_add_list:
		existing_item_codes = {item.get('item_code') for item in items}
		items_to_add = [code for code in force_add_list if code not in existing_item_codes]
		
		if items_to_add:
			# Get details of force add items
			force_add_items = frappe.db.sql("""
				SELECT 
					item_code,
					item_name,
					has_serial_no,
					brand,
					item_group
				FROM `tabItem`
				WHERE item_code IN ({items})
				AND disabled = 0
				ORDER BY item_code
			""".format(items=', '.join(['%s'] * len(items_to_add))), items_to_add, as_dict=1)
			
			items.extend(force_add_items)
	
	return items


def get_item_warehouse_data(item_code, warehouse_list, from_date, to_date, supplier_list, warehouse_hierarchy, warehouse_mapping):
	"""
	Get item data for each warehouse (for non-serialized items or items without supplier tracking)
	If warehouse is parent, aggregate data from children
	Uses bulk queries for better performance
	"""
	data = []
	processed_warehouses = set()
	
	# Get all warehouses including children for bulk query
	all_warehouses = set(warehouse_list)
	for parent, children in warehouse_hierarchy.items():
		all_warehouses.update(children)
	
	# Bulk query for stock balances
	stock_balances = get_bulk_stock_balance(item_code, list(all_warehouses))
	
	# Bulk query for sold quantities
	sold_quantities = get_bulk_sold_quantity(item_code, list(all_warehouses), from_date, to_date)
	
	for warehouse in warehouse_list:
		# Skip if already processed as a child
		if warehouse in processed_warehouses:
			continue
		
		# Check if this warehouse is a parent with children
		children = warehouse_hierarchy.get(warehouse, [])
		
		if children:
			# Parent warehouse - aggregate from children
			warehouses_to_aggregate = [warehouse] + children
			
			total_stock_balance = sum(stock_balances.get(wh, 0) for wh in warehouses_to_aggregate)
			total_sold_qty = sum(sold_quantities.get(wh, 0) for wh in warehouses_to_aggregate)
			
			# Mark children as processed
			for child_warehouse in children:
				processed_warehouses.add(child_warehouse)
			
			# Get supplier info if applicable
			supplier = None
			if supplier_list:
				supplier = get_item_supplier(item_code, supplier_list)
			
			# Use amplify code for warehouse display
			warehouse_code = warehouse_mapping.get(warehouse, warehouse)
			
			data.append({
				'item_code': item_code,
				'warehouse_code': warehouse_code,
				'sold_qty': total_sold_qty,
				'stock_balance': total_stock_balance,
				'supplier': supplier or ''
			})
			processed_warehouses.add(warehouse)
		else:
			# Child or standalone warehouse
			stock_balance = stock_balances.get(warehouse, 0)
			sold_qty = sold_quantities.get(warehouse, 0)
			
			# Get supplier info if applicable
			supplier = None
			if supplier_list:
				supplier = get_item_supplier(item_code, supplier_list)
			
			# Use amplify code for warehouse display
			warehouse_code = warehouse_mapping.get(warehouse, warehouse)
			
			data.append({
				'item_code': item_code,
				'warehouse_code': warehouse_code,
				'sold_qty': sold_qty,
				'stock_balance': stock_balance,
				'supplier': supplier or ''
			})
			processed_warehouses.add(warehouse)
	
	return data


def get_serial_number_data(item_code, warehouse_list, supplier_list, from_date, to_date, warehouse_hierarchy, warehouse_mapping):
	"""
	Get item data tracked by serial number and supplier
	Aggregates to parent warehouse if applicable
	Uses efficient bulk queries with serial number tracking
	"""
	# Get all warehouses including children
	all_warehouses = set(warehouse_list)
	for parent, children in warehouse_hierarchy.items():
		all_warehouses.update(children)
	
	all_warehouses_list = list(all_warehouses)
	
	# Get sold serial numbers from Stock Ledger Entry in date range
	sold_serials_query = """
		SELECT 
			sle.serial_no,
			sle.warehouse,
			sle.actual_qty,
			sle.posting_date
		FROM `tabStock Ledger Entry` sle
		WHERE sle.item_code = %s
		AND sle.posting_date BETWEEN %s AND %s
		AND sle.actual_qty < 0
		AND sle.voucher_type IN ('Delivery Note', 'Sales Invoice')
		AND sle.is_cancelled = 0
		AND sle.serial_no IS NOT NULL
		AND sle.serial_no != ''
		ORDER BY sle.posting_date
	"""
	
	sold_entries = frappe.db.sql(sold_serials_query, [item_code, from_date, to_date], as_dict=1)
	
	# Parse sold serial numbers and map to supplier
	sold_serials_map = {}  # serial_no -> {warehouse, qty, supplier}
	
	for entry in sold_entries:
		serial_nos_str = entry.get('serial_no', '')
		warehouse = entry.get('warehouse')
		qty = abs(entry.get('actual_qty', 0))
		
		# Parse serial numbers (can be comma or newline separated)
		serial_nos = [s.strip() for s in serial_nos_str.replace('\n', ',').split(',') if s.strip()]
		
		for serial_no in serial_nos:
			sold_serials_map[serial_no] = {
				'warehouse': warehouse,
				'qty': 1,  # Each serial = 1 unit
			}
	
	# Get all serial numbers with supplier info
	if sold_serials_map or all_warehouses_list:
		# Get current stock serials and sold serials
		serial_conditions = []
		serial_params = [item_code]
		
		# Add warehouse condition
		if all_warehouses_list:
			serial_conditions.append(f"sn.warehouse IN ({', '.join(['%s'] * len(all_warehouses_list))})")
			serial_params.extend(all_warehouses_list)
		
		# Add sold serials condition
		if sold_serials_map:
			sold_serial_list = list(sold_serials_map.keys())
			serial_conditions.append(f"sn.name IN ({', '.join(['%s'] * len(sold_serial_list))})")
			serial_params.extend(sold_serial_list)
		
		where_clause = " OR ".join(serial_conditions) if serial_conditions else "1=1"
		
		serial_nos_query = f"""
			SELECT 
				sn.name as serial_no,
				sn.warehouse,
				sn.supplier
			FROM `tabSerial No` sn
			WHERE sn.item_code = %s
			AND ({where_clause})
			ORDER BY sn.name
		"""
		
		serial_nos = frappe.db.sql(serial_nos_query, serial_params, as_dict=1)
	else:
		serial_nos = []
	
	# Build aggregated data by warehouse and supplier
	warehouse_supplier_data = {}
	
	for serial in serial_nos:
		serial_no = serial.get('serial_no')
		warehouse = serial.get('warehouse')
		supplier = serial.get('supplier')
		
		# Skip if supplier doesn't match filter
		if supplier_list and supplier not in supplier_list:
			continue
		
		# Check if this serial was sold
		sold_info = sold_serials_map.get(serial_no)
		sold_qty = sold_info['qty'] if sold_info else 0
		stock_balance = 1 if warehouse and not sold_info else 0
		
		# Determine reporting warehouse (parent if child)
		reporting_warehouse = get_reporting_warehouse(warehouse or (sold_info['warehouse'] if sold_info else None), warehouse_list, warehouse_hierarchy)
		warehouse_code = warehouse_mapping.get(reporting_warehouse, reporting_warehouse)
		
		# Aggregate by warehouse_code and supplier
		key = (warehouse_code, supplier or '')
		if key not in warehouse_supplier_data:
			warehouse_supplier_data[key] = {
				'item_code': item_code,
				'warehouse_code': warehouse_code,
				'sold_qty': 0,
				'stock_balance': 0,
				'supplier': supplier or ''
			}
		
		warehouse_supplier_data[key]['sold_qty'] += sold_qty
		warehouse_supplier_data[key]['stock_balance'] += stock_balance
	
	return list(warehouse_supplier_data.values())


def get_bulk_stock_balance(item_code, warehouse_list):
	"""
	Get current stock balance from Bin for multiple warehouses in one query
	Returns dict: {warehouse: actual_qty}
	"""
	if not warehouse_list:
		return {}
	
	result = frappe.db.sql("""
		SELECT warehouse, actual_qty
		FROM `tabBin`
		WHERE item_code = %s
		AND warehouse IN ({warehouses})
	""".format(warehouses=', '.join(['%s'] * len(warehouse_list))),
	[item_code] + warehouse_list, as_dict=1)
	
	return {row['warehouse']: flt(row['actual_qty']) for row in result}


def get_bulk_sold_quantity(item_code, warehouse_list, from_date, to_date):
	"""
	Calculate sold quantity from Stock Ledger Entry for multiple warehouses in one query
	Returns dict: {warehouse: sold_qty}
	Handles serial numbers properly by counting actual transaction quantities
	"""
	if not warehouse_list:
		return {}
	
	result = frappe.db.sql("""
		SELECT 
			warehouse,
			SUM(ABS(actual_qty)) as sold_qty
		FROM `tabStock Ledger Entry`
		WHERE item_code = %s
		AND warehouse IN ({warehouses})
		AND posting_date BETWEEN %s AND %s
		AND actual_qty < 0
		AND voucher_type IN ('Delivery Note', 'Sales Invoice')
		AND is_cancelled = 0
		GROUP BY warehouse
	""".format(warehouses=', '.join(['%s'] * len(warehouse_list))),
	[item_code] + warehouse_list + [from_date, to_date], as_dict=1)
	
	return {row['warehouse']: flt(row['sold_qty']) for row in result}


def get_item_supplier(item_code, supplier_list):
	"""
	Get the primary supplier for an item from recent purchases
	"""
	supplier = frappe.db.sql("""
		SELECT po.supplier
		FROM `tabPurchase Order Item` poi
		INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
		WHERE poi.item_code = %s
		AND po.docstatus = 1
		AND po.supplier IN ({suppliers})
		ORDER BY po.transaction_date DESC
		LIMIT 1
	""".format(suppliers=', '.join(['%s'] * len(supplier_list))), 
	[item_code] + supplier_list, as_dict=1)
	
	if not supplier:
		# Try Purchase Receipt
		supplier = frappe.db.sql("""
			SELECT pr.supplier
			FROM `tabPurchase Receipt Item` pri
			INNER JOIN `tabPurchase Receipt` pr ON pr.name = pri.parent
			WHERE pri.item_code = %s
			AND pr.docstatus = 1
			AND pr.supplier IN ({suppliers})
			ORDER BY pr.posting_date DESC
			LIMIT 1
		""".format(suppliers=', '.join(['%s'] * len(supplier_list))), 
		[item_code] + supplier_list, as_dict=1)
	
	return supplier[0].get('supplier') if supplier else None


def get_item_groups_with_children(item_group_list):
	"""
	Get all item groups including their children recursively
	"""
	if not item_group_list:
		return []
	
	all_item_groups = set(item_group_list)
	
	# Get all children recursively
	for item_group in item_group_list:
		children = frappe.db.sql("""
			WITH RECURSIVE item_group_tree AS (
				SELECT name, parent_item_group
				FROM `tabItem Group`
				WHERE parent_item_group = %s
				
				UNION ALL
				
				SELECT ig.name, ig.parent_item_group
				FROM `tabItem Group` ig
				INNER JOIN item_group_tree igt ON ig.parent_item_group = igt.name
			)
			SELECT name FROM item_group_tree
		""", [item_group], as_dict=1)
		
		for child in children:
			all_item_groups.add(child['name'])
	
	return list(all_item_groups)


def build_warehouse_hierarchy(warehouse_list):
	"""
	Build a mapping of parent warehouses to their children
	Only includes warehouses from the provided list
	"""
	hierarchy = {}
	
	# Get all warehouses with their parent info
	warehouse_parents = {}
	for warehouse in warehouse_list:
		parent = frappe.db.get_value('Warehouse', warehouse, 'parent_warehouse')
		warehouse_parents[warehouse] = parent
	
	# Build parent -> children mapping
	for warehouse in warehouse_list:
		parent = warehouse_parents.get(warehouse)
		
		# If this warehouse has a parent that's also in our list
		if parent and parent in warehouse_list:
			if parent not in hierarchy:
				hierarchy[parent] = []
			hierarchy[parent].append(warehouse)
		# If this warehouse has no parent, check if any warehouse is its child
		elif not parent:
			# Check if any warehouse in the list has this as parent
			children = [w for w in warehouse_list if warehouse_parents.get(w) == warehouse]
			if children:
				hierarchy[warehouse] = children
	
	return hierarchy


def get_reporting_warehouse(warehouse, warehouse_list, warehouse_hierarchy):
	"""
	Get the reporting warehouse (parent if exists in list, otherwise the warehouse itself)
	"""
	if not warehouse:
		return warehouse
	
	# Check if this warehouse has a parent in the warehouse_list
	parent = frappe.db.get_value('Warehouse', warehouse, 'parent_warehouse')
	
	if parent and parent in warehouse_list:
		return parent
	
	return warehouse


def generate_excel_report(data, from_date, to_date, reporter_id):
	"""
	Generate Excel file with the HP Amplify required format
	"""
	# Create workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "HP Amplify Report"
	
	# Define styles
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	header_font = Font(bold=True, color="FFFFFF", size=11)
	border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	# HP Amplify required headers
	headers = [
		'Reporter ID',
		'Partner Name',
		'Country',
		'Start period',
		'End period',
		'Transaction date',
		'HP Product Number',
		'Product EAN/UPC code',
		'Partner Product ID',
		'Inventory units',
		'Sales units',
		'Sell From Store ID',
		'Marketplace Name',
		'Store 1st line address',
		'Store City',
		'Store Post Code',
		'Store City',
		'Online Site Motion',
		'Product Origin',
		'Buying Price',
		'Street Price',
		'Currency code',
		'Partner Comment'
	]
	
	# Add headers
	for col_num, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col_num)
		cell.value = header
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center')
		cell.border = border
	
	# Add data
	row_num = 2
	for record in data:
		ws.cell(row=row_num, column=1, value=reporter_id)  # Reporter ID
		ws.cell(row=row_num, column=2, value='')  # Partner Name (empty)
		ws.cell(row=row_num, column=3, value='')  # Country (empty)
		ws.cell(row=row_num, column=4, value='')  # Start period (empty)
		ws.cell(row=row_num, column=5, value=str(to_date))  # End period
		ws.cell(row=row_num, column=6, value='')  # Transaction date (empty)
		ws.cell(row=row_num, column=7, value=record.get('item_code'))  # HP Product Number
		ws.cell(row=row_num, column=8, value='')  # Product EAN/UPC code (empty)
		ws.cell(row=row_num, column=9, value='')  # Partner Product ID (empty)
		ws.cell(row=row_num, column=10, value=record.get('stock_balance'))  # Inventory units
		ws.cell(row=row_num, column=11, value=record.get('sold_qty'))  # Sales units
		ws.cell(row=row_num, column=12, value=record.get('warehouse_code'))  # Sell From Store ID
		ws.cell(row=row_num, column=13, value='')  # Marketplace Name (empty)
		ws.cell(row=row_num, column=14, value='')  # Store 1st line address (empty)
		ws.cell(row=row_num, column=15, value='')  # Store City (empty)
		ws.cell(row=row_num, column=16, value='')  # Store Post Code (empty)
		ws.cell(row=row_num, column=17, value='')  # Store City (empty) - Second occurrence
		ws.cell(row=row_num, column=18, value='')  # Online Site Motion (empty)
		ws.cell(row=row_num, column=19, value='')  # Product Origin (empty)
		ws.cell(row=row_num, column=20, value='')  # Buying Price (empty)
		ws.cell(row=row_num, column=21, value='')  # Street Price (empty)
		ws.cell(row=row_num, column=22, value='')  # Currency code (empty)
		ws.cell(row=row_num, column=23, value='')  # Partner Comment (empty)
		
		# Apply borders
		for col_num in range(1, 24):
			ws.cell(row=row_num, column=col_num).border = border
		
		row_num += 1
	
	# Adjust column widths
	column_widths = [15, 15, 10, 12, 12, 15, 18, 18, 18, 15, 12, 18, 15, 20, 15, 12, 15, 18, 15, 12, 12, 12, 20]
	for i, width in enumerate(column_widths, 1):
		ws.column_dimensions[get_column_letter(i)].width = width
	
	# Save file with HP Amplify naming format
	from_date_str = from_date.strftime('%Y%m%d')
	to_date_str = to_date.strftime('%Y%m%d')
	filename = f"R1_POS_INV_AMPLIFY_{reporter_id}_{from_date_str}_{to_date_str}.xlsx"
	file_path = os.path.join(frappe.utils.get_site_path('private', 'files'), filename)
	wb.save(file_path)
	
	# Create File document
	file_doc = frappe.get_doc({
		'doctype': 'File',
		'file_name': filename,
		'is_private': 1,
		'file_url': f'/private/files/{filename}'
	})
	file_doc.insert(ignore_permissions=True)
	
	return file_doc.file_url
